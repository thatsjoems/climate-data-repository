"""
Generates the standardized Excel data-submission template - MODULE D.
Institutions download this template, fill it in, then upload it back into the system.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

TEMPLATE_VERSION = "v1.0"

REQUIRED_COLUMNS = [
    "loan_id", "borrower_name", "loan_amount_tzs", "collateral_type",
    "collateral_value_tzs", "region", "district", "reporting_period",
]
OPTIONAL_COLUMNS = ["climate_hazard_exposure"]

ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS

TANZANIA_REGIONS = [
    "Arusha", "Dar es Salaam", "Dodoma", "Geita", "Iringa", "Kagera", "Katavi",
    "Kigoma", "Kilimanjaro", "Lindi", "Manyara", "Mara", "Mbeya", "Morogoro",
    "Mtwara", "Mwanza", "Njombe", "Pwani", "Rukwa", "Ruvuma", "Shinyanga",
    "Simiyu", "Singida", "Songwe", "Tabora", "Tanga",
]

HAZARD_OPTIONS = ["None", "Drought", "Flood", "Cyclone", "Landslide"]


def generate_loan_collateral_template() -> bytes:
    """Generates an Excel (.xlsx) file with the standardized loan/collateral data layout."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan_Collateral_Data"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(ALL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 22

    # Example row showing how to fill in the template
    example = {
        "loan_id": "LN-2026-0001",
        "borrower_name": "Example Company Ltd",
        "loan_amount_tzs": 50000000,
        "collateral_type": "Land Title",
        "collateral_value_tzs": 80000000,
        "region": "Dodoma",
        "district": "Chamwino",
        "reporting_period": "2026-Q3",
        "climate_hazard_exposure": "Drought",
    }
    for col_idx, col_name in enumerate(ALL_COLUMNS, start=1):
        ws.cell(row=2, column=col_idx, value=example[col_name])

    # Dropdown validation for region and hazard, to help avoid free-text typing errors
    region_list = ",".join(TANZANIA_REGIONS)
    dv_region = DataValidation(type="list", formula1=f'"{region_list}"', allow_blank=False)
    ws.add_data_validation(dv_region)
    dv_region.add(f"{ws.cell(row=2, column=ALL_COLUMNS.index('region')+1).column_letter}2:"
                  f"{ws.cell(row=2, column=ALL_COLUMNS.index('region')+1).column_letter}1000")

    hazard_list = ",".join(HAZARD_OPTIONS)
    dv_hazard = DataValidation(type="list", formula1=f'"{hazard_list}"', allow_blank=True)
    ws.add_data_validation(dv_hazard)
    dv_hazard.add(f"{ws.cell(row=2, column=ALL_COLUMNS.index('climate_hazard_exposure')+1).column_letter}2:"
                  f"{ws.cell(row=2, column=ALL_COLUMNS.index('climate_hazard_exposure')+1).column_letter}1000")

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        [f"CDR Standardized Data Template - {TEMPLATE_VERSION}"],
        [""],
        ["1. Do not change the column header names (row 1)."],
        ["2. Delete the example row (row 2) before submitting your real data."],
        ["3. Columns marked as mandatory below must always be filled in."],
        ["4. reporting_period must follow the format: YYYY-Qn (e.g. 2026-Q3)."],
        ["5. loan_amount_tzs and collateral_value_tzs must be numbers only (no letters or symbols)."],
        ["6. region must be an actual Tanzanian region (choose from the dropdown)."],
        [""],
        ["Required columns: " + ", ".join(REQUIRED_COLUMNS)],
        ["Optional columns: " + ", ".join(OPTIONAL_COLUMNS)],
    ]
    for row in instructions:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 90

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
